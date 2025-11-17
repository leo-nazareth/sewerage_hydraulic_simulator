import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Carregar a planilha
file_path = '/home/ubuntu/upload/VazoeseCapacidadesRede.xlsx'

# Primeiro, vamos ver as abas disponíveis
wb = load_workbook(file_path, data_only=False)
print("Abas disponíveis:")
for sheet_name in wb.sheetnames:
    print(f"- {sheet_name}")

print("\n" + "="*50 + "\n")

# Analisar cada aba
for sheet_name in wb.sheetnames:
    print(f"ANÁLISE DA ABA: {sheet_name}")
    print("-" * 30)
    
    ws = wb[sheet_name]
    
    # Mostrar dimensões
    print(f"Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
    
    # Mostrar algumas células com valores e fórmulas
    print("\nCélulas com conteúdo (primeiras 20 linhas):")
    for row in range(1, min(21, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 10)):  # Limitar a 10 colunas
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {cell.coordinate}: FÓRMULA = {cell.value}")
                else:
                    print(f"  {cell.coordinate}: {cell.value}")
    
    print("\n" + "="*50 + "\n")

# Também vamos ler como DataFrame para ver os dados
print("DADOS COMO DATAFRAME:")
print("-" * 30)
try:
    df = pd.read_excel(file_path, sheet_name=None)
    for sheet_name, data in df.items():
        print(f"\nAba: {sheet_name}")
        print(f"Shape: {data.shape}")
        print("Primeiras linhas:")
        print(data.head())
        print("\n")
except Exception as e:
    print(f"Erro ao ler como DataFrame: {e}")

