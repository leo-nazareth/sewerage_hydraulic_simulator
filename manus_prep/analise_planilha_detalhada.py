# Análise mais detalhada da planilha Excel para entender as fórmulas corretas

import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Carregar a planilha e analisar as fórmulas específicas
file_path = '/home/ubuntu/upload/VazoeseCapacidadesRede.xlsx'
wb = load_workbook(file_path, data_only=False)

print("ANÁLISE DETALHADA DAS FÓRMULAS DA PLANILHA")
print("=" * 50)

# Analisar a aba CH que contém os cálculos hidráulicos
ws_ch = wb['CH']

print("\nFórmulas da aba CH:")
print("-" * 30)

# Mapear as células importantes
celulas_importantes = {
    'C6': 'Qcalc',
    'D6': 'Diametro', 
    'E6': 'Declividade',
    'F6': 'Manning',
    'G6': 'TETA',
    'H6': 'E',
    'I6': "E'",
    'J6': 'Iteracao1',
    'K6': 'Iteracao2',
    'L6': 'Iteracao3'
}

for coord, desc in celulas_importantes.items():
    cell = ws_ch[coord]
    if cell.value is not None:
        if isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"{desc} ({coord}): {cell.value}")
        else:
            print(f"{desc} ({coord}): {cell.value}")

# Analisar mais colunas para encontrar as iterações
print("\nAnalisando mais colunas para encontrar iterações:")
print("-" * 50)

for col in range(1, 33):  # Colunas A até AF
    for row in range(5, 7):  # Linhas 5 e 6
        cell = ws_ch.cell(row=row, column=col)
        if cell.value is not None:
            coord = cell.coordinate
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{coord}: {cell.value}")
            elif isinstance(cell.value, (int, float)) and cell.value != 0:
                print(f"{coord}: {cell.value}")

# Tentar ler os dados como DataFrame para ver os valores calculados
print("\n\nDADOS CALCULADOS DA ABA CH:")
print("-" * 30)

try:
    df_ch = pd.read_excel(file_path, sheet_name='CH', header=None)
    print("Dados da aba CH:")
    print(df_ch.iloc[4:7, 2:15])  # Linhas 5-7, colunas C-O
except Exception as e:
    print(f"Erro ao ler DataFrame: {e}")

# Analisar a aba principal também
print("\n\nANÁLISE DA ABA PRINCIPAL:")
print("-" * 30)

ws_main = wb['Vazoes e Calculo Hidraulico']

# Procurar por fórmulas relacionadas ao cálculo hidráulico
for row in range(20, 30):
    for col in range(1, 10):
        cell = ws_main.cell(row=row, column=col)
        if cell.value is not None:
            coord = cell.coordinate
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{coord}: {cell.value}")
            elif isinstance(cell.value, (int, float)) and abs(cell.value) > 0.001:
                print(f"{coord}: {cell.value}")

print("\n" + "=" * 50)

